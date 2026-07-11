#!/usr/bin/env python3
"""Health (QHP): real, live fetch from CMS's QHP Landscape file.

Source: data.healthcare.gov's DKAN metastore API -> a zip containing the
Individual Market Medical xlsx (PY2026, ~97K rows, the 30 states on the
federal marketplace). Produces qhp.json in the exact shape
ContentDatabase.swift/.kt already parse (county-keyed dict, plans grouped
by metal level, premiums keyed by age string) -- app-side parsing logic
does not need to change, only where the file comes from.
"""
import io
import json
import sys
import zipfile
from datetime import date, timezone, datetime

import openpyxl
import requests

DATASET_METASTORE_ID = "6fe7fb77-7291-4104-952f-7c7e2c5d0c45"  # QHP Landscape PY2026 Individual Medical
AGES = ["21", "27", "30", "40", "50", "60"]
AGE_COLUMNS = {  # 0-indexed column -> age band
    26: "21", 27: "27", 28: "30", 29: "40", 30: "50", 31: "60",
}
EXCLUDED_METAL_LEVELS = {"Catastrophic"}


def fetch_metadata():
    r = requests.get(
        f"https://data.healthcare.gov/api/1/metastore/schemas/dataset/items/{DATASET_METASTORE_ID}",
        timeout=30,
    )
    r.raise_for_status()
    meta = r.json()
    download_url = meta["distribution"][0]["downloadURL"]
    modified = meta.get("modified")
    return download_url, modified


def download_xlsx(download_url):
    r = requests.get(download_url, timeout=300)
    r.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise RuntimeError("no .xlsx file found inside the downloaded zip")
        return zf.read(xlsx_names[0])


def parse_price(cell):
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip().lstrip("$").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def build_counties(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    counties = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        state = row[0]
        if state is None:
            continue
        fips, county_name, metal, issuer, plan_id, plan_name = (
            row[1], row[2], row[3], row[4], row[6], row[7],
        )
        plan_type = row[9]
        if metal in EXCLUDED_METAL_LEVELS:
            continue
        premiums = {age: parse_price(row[col]) for col, age in AGE_COLUMNS.items()}
        premiums = {k: v for k, v in premiums.items() if v is not None}
        if len(premiums) != len(AGES):
            continue  # incomplete row -- skip rather than fabricate a missing age band

        key = f"{state}-{county_name}"
        county = counties.setdefault(key, {
            "state": state, "county": county_name, "fips": fips,
            "totalPlans": 0, "totalIssuers": 0, "plans": {}, "_issuers": set(),
        })
        county["plans"].setdefault(metal, []).append({
            "issuer": issuer, "name": plan_name, "type": plan_type, "p": premiums,
        })
        county["_issuers"].add(issuer)
        county["totalPlans"] += 1

    for county in counties.values():
        county["totalIssuers"] = len(county.pop("_issuers"))
    return counties


def main():
    download_url, modified = fetch_metadata()
    print(f"Downloading {download_url}", file=sys.stderr)
    xlsx_bytes = download_xlsx(download_url)
    counties = build_counties(xlsx_bytes)
    total_plans = sum(c["totalPlans"] for c in counties.values())
    print(f"Parsed {len(counties)} counties, {total_plans} plans", file=sys.stderr)

    snapshot_date = date.today().isoformat()
    with open("data/published/qhp.json", "w") as f:
        json.dump(counties, f, separators=(",", ":"))

    return {
        "dataset_id": "qhp",
        "domain_id": "health",
        "kind": "qhp_plans",
        "file_name": "qhp.json",
        "snapshot_date": snapshot_date,
        "source_kind": "auto",
        "source_name": "CMS QHP Landscape PY2026, Individual Market Medical -- full national file",
        "source_url": "https://data.healthcare.gov/qhp-landscape-files",
    }


if __name__ == "__main__":
    result = main()
    print(json.dumps(result))
